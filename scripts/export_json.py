"""Export apartment_database.xlsx to database.json for the viewer.

Run from the project root:
    python scripts/export_json.py

The viewer (viewer.html) loads database.json on page load. Re-run this script
after any edit to database.xlsx.
"""
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Missing openpyxl. Install with: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "database.xlsx"
JSON_OUT = ROOT / "database.json"
THUMBS = ROOT / "thumbnails"

def main():
    if not XLSX.exists():
        sys.exit(f"Not found: {XLSX}")

    wb = load_workbook(XLSX)
    ws = wb["Apartments"]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]:
            continue
        # Column order: A=thumbnail (image, not a value), B=ID, C=beds, D=baths,
        # E=frontage, F=width, G=width2, H=depth, I=ref, J=area, K=study
        uid = r[1]
        thumb_path = THUMBS / f"{uid}.png"
        rows.append({
            "id": uid,
            "bedrooms": r[2],
            "bathrooms": r[3],
            "frontage": r[4],
            "width": r[5],
            "width2": r[6],
            "depth": r[7],
            "reference": r[8] or "",
            "area": r[9],
            "study": r[10] == "Yes" if r[10] else False,
            "thumbnail": f"thumbnails/{uid}.png" if thumb_path.exists() else None,
        })

    JSON_OUT.write_text(json.dumps(rows, indent=2))
    print(f"Exported {len(rows)} units to {JSON_OUT}")

if __name__ == "__main__":
    main()
